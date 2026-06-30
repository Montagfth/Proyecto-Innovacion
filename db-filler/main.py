import os
import sys
from datetime import datetime, timedelta
import datetime as dt
import openpyxl
from dotenv import load_dotenv
import libsql_client

# Import the price calculator from the current directory
from price_calculator import PrintOrder, calculate_print_price

# Load environment variables
load_dotenv()

TURSO_DATABASE_URL = os.getenv("TURSO_DATABASE_URL", "file:local.db")
TURSO_AUTH_TOKEN = os.getenv("TURSO_AUTH_TOKEN", "")

# Fallback to local DB for local testing if the URL is the default example one
if "your-database-name.turso.io" in TURSO_DATABASE_URL or not TURSO_DATABASE_URL:
    print("Detected example or empty TURSO_DATABASE_URL. Falling back to local database 'file:local.db'.")
    TURSO_DATABASE_URL = "file:local.db"

def normalize_str(s):
    """Normalize string to facilitate matching: trim, lowercase, remove accents and ñ."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    s = s.replace("ñ", "n")
    return s

def get_date_by_weekday(day_name):
    """
    Map weekdays to dates in the current week starting Mon Jun 22 to Sun Jun 28, 2026.
    Allows matching DiaSemana from Excel to actual dates.
    """
    day_norm = normalize_str(day_name)
    mapping = {
        "lunes": "2026-06-22",
        "martes": "2026-06-23",
        "miercoles": "2026-06-24",
        "jueves": "2026-06-25",
        "viernes": "2026-06-26",
        "sabado": "2026-06-27",
        "domingo": "2026-06-28"
    }
    return mapping.get(day_norm, "2026-06-24")

def parse_time_str(val):
    """Parse time cell values which can be string, datetime.time, or datetime.datetime."""
    if isinstance(val, dt.time):
        return val.strftime("%H:%M")
    elif isinstance(val, dt.datetime):
        return val.strftime("%H:%M")
    elif val is not None:
        return str(val).strip()
    return "00:00"

def initialize_database_if_empty(client):
    """Initialize database tables using db.sql if print_types table does not exist or is empty."""
    try:
        # Check if print_types exists
        res = client.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='print_types'")
        if len(res.rows) > 0:
            # Table exists, check if it has entries
            count_res = client.execute("SELECT COUNT(*) FROM print_types")
            if count_res.rows[0][0] > 0:
                print("Database tables already initialized.")
                return
        
        print("Database is empty or unitialized. Initializing from db.sql...")
        
        # Locate db.sql
        possible_paths = [
            os.path.join(os.path.dirname(__file__), "..", "db.sql"),
            os.path.join(os.path.dirname(__file__), "db.sql"),
            "db.sql"
        ]
        schema_path = None
        for path in possible_paths:
            if os.path.exists(path):
                schema_path = path
                break
        
        if not schema_path:
            print("Warning: db.sql file not found. Skipping schema initialization.")
            return

        with open(schema_path, "r", encoding="utf-8") as f:
            sql_content = f.read()

        # Split SQL content into statements
        statements = []
        current_stmt = []
        for line in sql_content.splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("--") or stripped.startswith("---"):
                continue
            current_stmt.append(line)
            if stripped.endswith(";"):
                statements.append("\n".join(current_stmt))
                current_stmt = []

        print(f"Executing {len(statements)} schema statements...")
        for stmt in statements:
            stmt_str = stmt.strip()
            if stmt_str:
                client.execute(stmt_str)
        print("Database schema successfully initialized.")
    except Exception as e:
        print(f"Error during schema initialization: {e}")

def main():
    print("Connecting to database at:", TURSO_DATABASE_URL)
    
    # Establish connection
    try:
        if TURSO_DATABASE_URL.startswith("file:") or "localhost" in TURSO_DATABASE_URL or "127.0.0.1" in TURSO_DATABASE_URL:
            # Local connection (usually doesn't need auth token)
            client = libsql_client.create_client_sync(TURSO_DATABASE_URL)
        else:
            # Remote Turso connection
            client = libsql_client.create_client_sync(TURSO_DATABASE_URL, auth_token=TURSO_AUTH_TOKEN)
    except Exception as conn_err:
        print(f"Fatal error connecting to database: {conn_err}")
        sys.exit(1)

    try:
        # Initialize schema if database is empty/fresh
        initialize_database_if_empty(client)

        # Retrieve dynamic mappings from DB tables
        print("Fetching lookup tables from database...")
        
        res_types = client.execute("SELECT type_id, print_type FROM print_types")
        print_types_map = {normalize_str(row[1]): row[0] for row in res_types.rows}
        
        res_sizes = client.execute("SELECT size_id, print_size FROM print_sizes")
        print_sizes_map = {normalize_str(row[1]): row[0] for row in res_sizes.rows}
        
        res_materials = client.execute("SELECT material_id, print_material FROM print_materials")
        print_materials_map = {normalize_str(row[1]): row[0] for row in res_materials.rows}
        
        res_machines = client.execute("SELECT machine_id, print_machine FROM print_machines")
        print_machines_map = {normalize_str(row[1]): row[0] for row in res_machines.rows}

        print(f"Loaded mappings: Types: {len(print_types_map)}, Sizes: {len(print_sizes_map)}, Materials: {len(print_materials_map)}, Machines: {len(print_machines_map)}")

        # Load Excel dataset
        excel_path = "dataset.xlsx"
        if not os.path.exists(excel_path):
            excel_path = os.path.join(os.path.dirname(__file__), "dataset.xlsx")
        
        if not os.path.exists(excel_path):
            print(f"Fatal: Excel dataset file '{excel_path}' not found.")
            sys.exit(1)
            
        print(f"Loading Excel workbook from: {excel_path}...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        print(f"Active sheet: {sheet.title}, Max rows: {sheet.max_row}")

        # Get header names
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 13)]
        header_to_col = {}
        for i, h in enumerate(headers):
            if h:
                header_to_col[normalize_str(h)] = i + 1

        print("Dataset columns mapped:", list(header_to_col.keys()))

        def get_cell_val(row_idx, col_name):
            col_idx = header_to_col.get(normalize_str(col_name))
            if col_idx:
                return sheet.cell(row=row_idx, column=col_idx).value
            return None

        # Start insertion transaction loop
        inserted_count = 0
        error_count = 0

        # We will insert records inside database
        for r in range(2, sheet.max_row + 1):
            id_pedido = get_cell_val(r, "ID_Pedido")
            if id_pedido is None:
                # Skip empty row
                continue

            tipo_trabajo = get_cell_val(r, "TipoTrabajo")
            cantidad = get_cell_val(r, "Cantidad")
            tamano = get_cell_val(r, "Tamaño")
            material = get_cell_val(r, "Material")
            color = get_cell_val(r, "Color")
            prioridad = get_cell_val(r, "Prioridad")
            maquina = get_cell_val(r, "MaquinaAsignada")
            hora_ingreso = get_cell_val(r, "HoraIngreso")
            dia_semana = get_cell_val(r, "DiaSemana")
            tiempo_prod = get_cell_val(r, "TiempoProduccionMinutos")

            # Validate categorical matches
            norm_type = normalize_str(tipo_trabajo)
            norm_size = normalize_str(tamano)
            norm_material = normalize_str(material)
            norm_machine = normalize_str(maquina)

            db_type_id = print_types_map.get(norm_type)
            db_size_id = print_sizes_map.get(norm_size)
            db_material_id = print_materials_map.get(norm_material)
            db_machine_id = print_machines_map.get(norm_machine)

            missing_lookups = []
            if db_type_id is None: missing_lookups.append(f"type '{tipo_trabajo}'")
            if db_size_id is None: missing_lookups.append(f"size '{tamano}'")
            if db_material_id is None: missing_lookups.append(f"material '{material}'")
            if db_machine_id is None: missing_lookups.append(f"machine '{maquina}'")

            if missing_lookups:
                print(f"Skipping row {r} (ID {id_pedido}) due to missing database mappings: {', '.join(missing_lookups)}")
                error_count += 1
                continue

            # Convert color and priority
            color_bool = bool(int(color)) if color is not None else False
            priority_val = int(prioridad) if prioridad is not None else 0

            # Calculate Price
            try:
                order_obj = PrintOrder(
                    job_type=tipo_trabajo,
                    quantity=int(cantidad),
                    size=tamano,
                    material=material,
                    color=color_bool
                )
                total_price = calculate_print_price(order_obj)
                total_int = int(round(total_price))
            except Exception as calc_err:
                print(f"Warning: Price calculation failed for row {r} (ID {id_pedido}): {calc_err}. Using 0.")
                total_int = 0

            # Calculate timestamps
            date_str = get_date_by_weekday(dia_semana)
            time_str = parse_time_str(hora_ingreso)

            try:
                start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            except Exception as parse_err:
                print(f"Warning: Could not parse time '{date_str} {time_str}' on row {r}: {parse_err}. Using default.")
                start_dt = datetime.now()

            # Add production minutes
            prod_minutes = float(tiempo_prod) if tiempo_prod is not None else 0.0
            completed_dt = start_dt + timedelta(minutes=prod_minutes)

            created_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
            started_at_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
            completed_at_str = completed_dt.strftime("%Y-%m-%d %H:%M:%S")

            estimated_time_int = int(round(prod_minutes))
            final_time_int = int(round(prod_minutes))

            # Status is completed
            status_val = "Completado"
            client_name = "Cliente Importado"

            # Execute INSERT OR REPLACE
            try:
                client.execute(
                    """
                    INSERT OR REPLACE INTO orders (
                        order_id, client, print_type, print_size, print_material, print_machine,
                        quantity, total, priority, created, status, started_at, completed_at,
                        estimated_time, final_time
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        int(id_pedido),
                        client_name,
                        db_type_id,
                        db_size_id,
                        db_material_id,
                        db_machine_id,
                        int(cantidad),
                        total_int,
                        priority_val,
                        created_str,
                        status_val,
                        started_at_str,
                        completed_at_str,
                        estimated_time_int,
                        final_time_int
                    ]
                )
                inserted_count += 1
            except Exception as insert_err:
                print(f"Error inserting row {r} (ID {id_pedido}): {insert_err}")
                error_count += 1

        print(f"Filler execution finished. Total successfully inserted/replaced: {inserted_count}, Errors: {error_count}")
    finally:
        client.close()

if __name__ == "__main__":
    main()
